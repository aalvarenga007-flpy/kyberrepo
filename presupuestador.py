# -*- coding: utf-8 -*-
"""
PRESUPUESTADOR - Conepasa IA / Ejapo Comercial San Jose
=======================================================

Lee la lista de provista que manda un cliente (Excel, CSV, PDF o foto),
la resuelve contra el catalogo y genera un presupuesto para revisar.

USO DESDE LA CONSOLA:

    python presupuestador.py lista_del_cliente.xlsx
    python presupuestador.py lista.pdf --ruc 93181-0
    python presupuestador.py foto.jpg --ruc 80058190-3 --referencia EMPEDRIL

QUE HACE:
    1. Extrae las lineas del archivo (que puede venir de cualquier forma)
    2. Para cada linea busca el producto en el catalogo, con 3 niveles:
         ALTA  - el cliente ya lo compro y siempre el mismo -> precargado
         MEDIA - el cliente lo compro pero alterna marcas   -> 3 opciones
         BAJA  - cliente nuevo, usa el ranking general      -> 3 opciones
    3. Convierte kilos/litros/cajas a la cantidad de bultos que corresponde
    4. Genera un Excel para revisar antes de emitir

NUNCA emite solo. Siempre pasa por la pantalla de revision.
"""

import os
import re
import sys
import unicodedata
from datetime import datetime

import pandas as pd
import pymysql
from dotenv import load_dotenv

load_dotenv()

# ---------------------------------------------------------------- conexion

DB = dict(
    host=os.getenv("MYSQL_HOST", "localhost"),
    user=os.getenv("MYSQL_USER"),
    password=os.getenv("MYSQL_PASSWORD"),
    database=os.getenv("MYSQL_DATABASE", "ejapo_sanjose_bi"),
    charset="utf8mb4",
    cursorclass=pymysql.cursors.DictCursor,
)

# Carpeta donde viven los archivos de presupuestos, fuera de claude_engine.
# Se puede cambiar con PRESUPUESTOS_DIR en el .env
BASE_DIR = os.getenv(
    "PRESUPUESTOS_DIR",
    os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "presupuestos"))
DIR_ENTRADA = os.path.join(BASE_DIR, "entrada")
DIR_SALIDA = os.path.join(BASE_DIR, "salida")

for _d in (DIR_ENTRADA, DIR_SALIDA):
    os.makedirs(_d, exist_ok=True)

# Paginas de PDF escaneado que se mandan al modelo. Mas que esto encarece la
# lectura sin aportar: los pedidos rara vez pasan de 6 hojas.
MAX_PAGINAS_PDF = 6

# Diagnostico de la ultima lectura, para que la pantalla pueda mostrar que
# pestañas se leyeron y cuales fallaron en vez de perderlas en silencio.
ULTIMA_LECTURA = {"hojas": [], "errores": []}

# umbral de concentracion para considerar una eleccion "clara"
UMBRAL_ALTA = 0.70

# meses de historial del cliente que se miran
MESES_HISTORIAL = 12


def conectar():
    return pymysql.connect(**DB)


def consultar(sql, params=None):
    with conectar() as cn, cn.cursor() as cur:
        cur.execute(sql, params or ())
        return cur.fetchall()


# ---------------------------------------------------------------- texto

def norm(s):
    """Mayusculas, sin acentos, sin espacios de mas."""
    s = unicodedata.normalize("NFD", str(s))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", s.upper().strip())


UNIDADES = {
    "LITROS": "L", "LITRO": "L", "LTS": "L", "LT": "L", "L": "L",
    "KILOS": "KG", "KILO": "KG", "KG": "KG", "KGS": "KG", "K": "KG",
    "UNIDADES": "U", "UNIDAD": "U", "UNIDS": "U", "UND": "U", "U": "U",
    "GRAMOS": "GR", "GR": "GR", "G": "GR", "ML": "ML", "CC": "CC",
}

# unidades que significan "un bulto", no una medida
BULTOS = {"CAJA", "CAJAS", "BOLSA", "BOLSAS", "FARDO", "FARDOS",
          "PLANCHA", "PLANCHAS", "TIRA", "TIRAS", "PAQUETE", "PAQUETES",
          "BIDON", "BIDONES", "PAQ", "FDO"}

# Bultos ambiguos: 'paquete' puede ser el envase del producto (1 paquete de
# fosforos = 1 unidad) o un fardo. Sin equivalencia declarada NO se multiplica:
# cotizar de mas es peor que cotizar de menos.
BULTOS_AMBIGUOS = {"PAQUETE", "PAQ", "BIDON", "UNIDAD"}

# Palabras que aparecen en los pedidos pero no identifican un producto: no
# sirven para decidir marca ni variedad.
PALABRAS_VACIAS = {
    "PARA", "CON", "SIN", "DE", "DEL", "LOS", "LAS", "UNA", "UNO", "POR",
    "TIPO", "GRANDE", "CHICO", "CHICA", "MEDIANO", "MEDIANA", "COMUN",
    "VARIEDAD", "VARIEDADES", "VARIOS", "VARIAS", "SURTIDO", "SURTIDA",
    "UNIDAD", "UNIDADES", "PAQUETE", "PAQUETES", "CAJA", "CAJAS", "BOLSA",
    "BOLSAS", "KILO", "KILOS", "LITRO", "LITROS", "SABOR", "SABORES",
    "ENTERA", "ENTERO", "PRIMERA", "SEGUNDA", "MEJOR", "SIMILAR", "OTRO",
}


def parse_cantidad(txt):
    """
    '10 litros' -> (10.0, 'L')   '1 caja' -> (1.0, 'CAJA')   '-' -> (None, None)
    '1/2 bolsa' -> (0.5, 'BOLSA')

    Las fracciones aparecen seguido en los pedidos escritos a mano
    ('1/2 kg de levadura', 'coquito 1/2 bolsa') y sin esto se leian como 1.
    """
    t = norm(txt)
    if not t or t in ("-", "--", "S/D", "NAN", "X"):
        return (None, None)

    # fraccion: 1/2, 3/4, y tambien mixtas como 1 1/2
    m = re.match(r"(?:(\d+)\s+)?(\d+)\s*/\s*(\d+)\s*(.*)", t)
    if m and int(m.group(3)) != 0:
        entero = float(m.group(1) or 0)
        qty = entero + float(m.group(2)) / float(m.group(3))
        resto = m.group(4).strip().rstrip(".")
    else:
        m = re.match(r"([\d]+(?:[.,]\d+)?)\s*(.*)", t)
        if not m:
            return (None, t)
        qty = float(m.group(1).replace(",", "."))
        resto = m.group(2).strip().rstrip(".")

    if resto in BULTOS:
        return (qty, resto.rstrip("S") if resto not in ("PAQ", "FDO") else resto)
    return (qty, UNIDADES.get(resto, resto))


def tamano_producto(descripcion):
    """
    'SM AZUCAR 5KG' -> (5.0, 'KG').  'LEVADURA DE 500GRS' -> (500.0, 'GR').

    Se usa tanto sobre el catalogo como sobre el texto del cliente, asi que
    tiene que tolerar todas las formas de escribir lo mismo: 5KG, 5 KG, 5K,
    5 KILOS, 500GR, 500GRS, 500 GS, 1LT, 1 LTS.
    """
    d = norm(descripcion)
    # La unidad puede venir pegada al numero ("500GRS") o separada ("5 KILOS"),
    # asi que se normaliza junto con el numero. Buscar la unidad sola no sirve:
    # en "500GRS" no hay separacion entre el 0 y la G.
    d = re.sub(r"(\d+(?:[.,]\d+)?)\s*(?:KILOS?|KGS?|KG|K)\b", r"\1 KG", d)
    d = re.sub(r"(\d+(?:[.,]\d+)?)\s*(?:GRAMOS?|GRS|GR|GS|G)\b", r"\1 GR", d)
    d = re.sub(r"(\d+(?:[.,]\d+)?)\s*(?:LITROS?|LTS|LT|L)\b", r"\1 L", d)
    d = re.sub(r"(\d+(?:[.,]\d+)?)\s*(ML|CC)\b", r"\1 \2", d)
    m = re.search(r"(\d+(?:[.,]\d+)?)\s+(KG|L|GR|ML|CC)\b", d)
    if not m:
        return (None, None)
    return (float(m.group(1).replace(",", ".")), m.group(2))


def _mismo_tamano(a, b):
    """Compara dos tamanos tolerando KG/GR y L/ML."""
    if not a[0] or not b[0]:
        return False
    if a[1] == b[1]:
        return abs(a[0] - b[0]) < 0.01
    equivalencias = {("KG", "GR"): 1000, ("L", "ML"): 1000, ("L", "CC"): 1000}
    for (grande, chico), factor in equivalencias.items():
        if a[1] == grande and b[1] == chico:
            return abs(a[0] * factor - b[0]) < 1
        if a[1] == chico and b[1] == grande:
            return abs(a[0] - b[0] * factor) < 1
    return False


# ---------------------------------------------------------------- lectura de archivos

def leer_archivo(ruta):
    """
    Extrae las lineas del pedido, sea cual sea el formato.

    Los clientes mandan de todo: columnas al reves, la cantidad metida dentro
    del texto ("45 k arroz"), varios productos en un renglon, fracciones,
    secciones con totales, y una pestaña o columna por destino. Escribir reglas
    para cada caso es una carrera perdida, asi que la lectura la hace el
    modelo, que entiende el texto igual que una persona.

    Un Excel con varias pestañas se procesa de a UNA pestaña por llamada. Todo
    junto en una sola llamada hacia que la respuesta se cortara por la mitad y
    el pedido se leyera incompleto.

    Devuelve: [{'texto':..., 'cantidad_txt':..., 'destino':...}]
    """
    if not os.getenv("ANTHROPIC_API_KEY"):
        return _leer_por_columnas(ruta)

    partes = _partir(ruta)          # [(nombre_o_None, texto, imagenes)]
    filas = []
    errores = []
    resumen = []
    ULTIMA_LECTURA["hojas"] = []
    ULTIMA_LECTURA["errores"] = []
    for nombre, contenido, imagenes in partes:
        try:
            nuevas = _extraer_con_modelo(contenido, imagenes, nombre)
        except Exception as error:  # noqa: BLE001
            # Una hoja larga puede cortar la respuesta. Se reintenta partida
            # al medio antes de darla por perdida: perder una hoja entera es
            # mucho peor que hacer dos llamadas.
            nuevas = []
            if contenido and len(contenido) > 2000:
                renglones = contenido.split("\n")
                mitad = len(renglones) // 2
                for trozo in (renglones[:mitad], renglones[mitad:]):
                    try:
                        nuevas += _extraer_con_modelo("\n".join(trozo), None, nombre)
                    except Exception:  # noqa: BLE001
                        pass
            if not nuevas:
                errores.append(f"{nombre or 'archivo'}: {error}")
                continue
            print(f"  hoja '{nombre}': se leyo en dos partes")
        # El destino final combina la pestaña con la seccion de adentro, porque
        # los dos niveles importan: la hoja "Carla Maria" tiene adentro
        # "Viveres Casino" y tres residencias, y cada una se cotiza aparte.
        hoja_es_destino = bool(nombre) and len(partes) > 1 and not _es_rubro(nombre)
        for f in nuevas:
            seccion = (f.get("destino") or "").strip()
            if seccion and _es_rubro(seccion):
                seccion = ""          # "LIMPIEZA" sola no es un destino
            if hoja_es_destino and seccion:
                f["destino"] = f"{nombre.strip()} · {seccion}"
            elif hoja_es_destino:
                f["destino"] = nombre.strip()
            else:
                f["destino"] = seccion
        filas += nuevas
        if nombre:
            resumen.append((nombre, len(nuevas)))
            print(f"  hoja '{nombre}': {len(nuevas)} lineas")

    if filas:
        if errores:
            print(f"  ojo: {len(errores)} parte(s) no se pudieron leer -> "
                  + " ; ".join(errores[:3]))
        ULTIMA_LECTURA["hojas"] = resumen
        ULTIMA_LECTURA["errores"] = errores
        return filas

    detalle = " ; ".join(errores[:3]) if errores else "sin resultado"
    print(f"  (la lectura asistida fallo: {detalle} - se usan las reglas)")
    if any(img for _, _, img in partes):
        raise RuntimeError(
            f"No se pudo leer el archivo. Detalle: {detalle}")
    return _leer_por_columnas(ruta)


def _partir(ruta):
    """
    Parte el archivo en pedazos manejables para el modelo.
    Excel -> una parte por pestaña.  PDF/imagen -> una sola parte.
    """
    ext = os.path.splitext(ruta)[1].lower()
    if ext in (".xlsx", ".xls", ".xlsm"):
        partes = []
        hojas = pd.read_excel(ruta, sheet_name=None, header=None, dtype=str)
        for nombre, hoja in hojas.items():
            hoja = hoja.dropna(how="all").dropna(axis=1, how="all")
            if hoja.empty:
                continue
            texto = hoja.fillna("").astype(str).to_csv(sep="|", index=False, header=False)
            partes.append((str(nombre).strip(), texto, None))
        if len(partes) > 1:
            print(f"  el Excel tiene {len(partes)} pestañas con datos")
        return partes or [(None, "", None)]

    contenido, imagenes = _preparar(ruta)
    return [(None, contenido, imagenes)]


def _preparar(ruta):
    """Devuelve (texto_del_archivo, dict_imagen_o_None) para mandarle al modelo."""
    ext = os.path.splitext(ruta)[1].lower()

    if ext in (".jpg", ".jpeg", ".png", ".webp"):
        import base64
        mime = {"png": "image/png", "webp": "image/webp"}.get(ext.lstrip("."), "image/jpeg")
        with open(ruta, "rb") as f:
            return "", [{"media_type": mime,
                         "data": base64.b64encode(f.read()).decode()}]

    if ext in (".xlsx", ".xls", ".xlsm"):
        partes = []
        hojas = pd.read_excel(ruta, sheet_name=None, header=None, dtype=str)
        con_datos = [n for n, h in hojas.items() if not h.dropna(how="all").empty]
        if len(con_datos) > 1:
            partes.append(
                f"NOTA: este Excel tiene {len(con_datos)} pestañas: "
                f"{', '.join(con_datos)}. Fijate si son destinos distintos "
                f"(estancias, retiros) o rubros del mismo pedido.")
        for nombre, hoja in hojas.items():
            hoja = hoja.dropna(how="all").dropna(axis=1, how="all")
            if hoja.empty:
                continue
            partes.append(f"--- HOJA: {nombre} ---")
            partes.append(hoja.fillna("").astype(str).to_csv(sep="|", index=False, header=False))
        return "\n".join(partes), None

    if ext in (".csv", ".tsv"):
        with open(ruta, encoding="utf-8-sig", errors="replace") as f:
            return f.read(), None

    if ext == ".pdf":
        import pdfplumber
        with pdfplumber.open(ruta) as pdf:
            texto = "\n".join((p.extract_text() or "") for p in pdf.pages)
        if len(texto.strip()) > 80:
            return texto, None

        # PDF escaneado: van TODAS las paginas como imagen. Mandar solo la
        # primera hacia que un pedido de 4 hojas se leyera hasta la mitad de
        # la primera y el resto desapareciera sin aviso.
        import base64, io
        imagenes = []
        with pdfplumber.open(ruta) as pdf:
            for pagina in pdf.pages[:MAX_PAGINAS_PDF]:
                buf = io.BytesIO()
                pagina.to_image(resolution=140).original.save(
                    buf, format="JPEG", quality=80)
                imagenes.append({"media_type": "image/jpeg",
                                 "data": base64.b64encode(buf.getvalue()).decode()})
            if len(pdf.pages) > MAX_PAGINAS_PDF:
                print(f"  ojo: el PDF tiene {len(pdf.pages)} paginas y se leen "
                      f"las primeras {MAX_PAGINAS_PDF}")
        return "", imagenes

    raise ValueError(f"No se como leer un archivo {ext}")


PROMPT_EXTRACCION = """Sos un asistente de una distribuidora mayorista en Paraguay.
Te paso el pedido de provista que mando un cliente. Extrae TODAS las lineas de
productos.

Devolve UNICAMENTE un JSON, sin explicaciones y sin backticks:
[{"texto": "nombre del producto", "cantidad_txt": "30 Kg", "destino": ""}]

Reglas:
- "texto": el producto tal como lo escribio el cliente, sin la cantidad adentro.
  Si escribio la marca o el tamano, dejalos ("Arroz Tio Nico 5kg" queda entero).
- "cantidad_txt": el numero con su unidad tal cual ("30 Kg", "1 caja",
  "3 planchas", "1/2 bolsa", "10 lts"). Respeta las fracciones.
- Si la cantidad esta metida dentro del texto ("45 k arroz", "Huevo 24 planchas",
  "1 bolsa kokito"), separala: texto="arroz", cantidad_txt="45 k".
- Si un renglon tiene VARIOS productos ("2 paq anis, comino, pimienta" o
  "FIDEOS: tallarin, monito, caracolito"), devolve una linea por cada producto,
  repartiendo o repitiendo la cantidad segun corresponda.
- DESTINOS. Un mismo pedido puede ir a lugares distintos, y eso puede venir
  de dos formas:
  a) en COLUMNAS de cantidad (ADM, RETIRO 1, RETIRO 2, CENTRAL, sucursales)
  b) en PESTAÑAS distintas del Excel (aparecen como "--- HOJA: nombre ---")
  En los dos casos devolve una linea por cada destino con cantidad y poné el
  nombre del destino en "destino".
- Una misma pestaña puede tener VARIAS SECCIONES que son destinos distintos:
  encabezados en medio de la planilla como "VIVERES CASINO", "LIMPIEZA
  RESIDENCIA 1", "LIMPIEZA RESIDENCIA 2", "HARAS", "SOJALES", "GUARDIAS",
  con sus productos debajo. Cuando pase eso, poné ese encabezado en "destino"
  para cada producto que le corresponda.
- OJO con las pestañas: el nombre de una pestaña es un destino solo si nombra
  un LUGAR (una estancia, un retiro, una sucursal, un casco, una seccion de
  campo). Si nombra un RUBRO o categoria de mercaderia (VIVERES, ALMACEN,
  LIMPIEZA, CARNE, VERDULERIA, BEBIDAS, PANADERIA, FRUTAS) NO es un destino:
  en ese caso dejá "destino" en "" y devolve todo junto, porque es un solo
  pedido dividido por rubro.
  Si dudás entre las dos cosas, dejá "destino" en "".
- Si hay un solo destino, dejá "destino" en "".
- IGNORA: encabezados, titulos de seccion (Almacen, Carne, Limpieza, Frutas),
  filas de total o subtotal, columnas de precio, y las lineas sin cantidad o
  con guion (esas el cliente no las quiere).
- NO inventes productos que no esten en el pedido. NO corrijas la ortografia.

Pedido:
"""


def _extraer_con_modelo(contenido, imagen, nombre_hoja=None):
    import json
    import urllib.request

    bloques = []
    for img in (imagen or []):
        bloques.append({"type": "image", "source": {"type": "base64", **img}})
    if imagen and len(imagen) > 1:
        bloques.append({"type": "text",
                        "text": f"Son {len(imagen)} paginas del MISMO pedido. "
                                "Extrae los productos de TODAS."})
    encabezado = PROMPT_EXTRACCION
    if nombre_hoja:
        encabezado += f"\n(Esto es la pestaña \"{nombre_hoja}\" del Excel.)\n"
    bloques.append({"type": "text",
                    "text": encabezado + (contenido[:120000] if contenido else "")})

    cuerpo = json.dumps({
        "model": os.getenv("ANTHROPIC_MODEL_EXTRACCION", "claude-sonnet-5"),
        "max_tokens": 16000,
        "messages": [{"role": "user", "content": bloques}],
    }).encode()

    pedido = urllib.request.Request(
        "https://api.anthropic.com/v1/messages",
        data=cuerpo,
        headers={"content-type": "application/json",
                 "x-api-key": os.getenv("ANTHROPIC_API_KEY"),
                 "anthropic-version": "2023-06-01"},
    )
    with urllib.request.urlopen(pedido, timeout=180) as r:
        data = json.loads(r.read())

    texto = "".join(b.get("text", "") for b in data.get("content", []))
    texto = texto.replace("```json", "").replace("```", "").strip()
    inicio, fin = texto.find("["), texto.rfind("]")
    if inicio < 0:
        raise ValueError("el modelo no devolvio un JSON")
    if fin < inicio:
        # respuesta cortada: se recupera hasta el ultimo objeto completo
        corte = texto.rfind("}")
        if corte < 0:
            raise ValueError("respuesta cortada y sin datos recuperables")
        texto = texto[:corte + 1] + "]"
        fin = len(texto) - 1
    try:
        filas = json.loads(texto[inicio:fin + 1])
    except json.JSONDecodeError:
        corte = texto.rfind("},")
        if corte < 0:
            raise
        filas = json.loads(texto[inicio:corte + 1] + "]")
    limpias = []
    for f in filas:
        if not isinstance(f, dict):
            continue
        texto_prod = str(f.get("texto", "")).strip()
        if len(texto_prod) < 2:
            continue
        limpias.append({
            "texto": texto_prod,
            "cantidad_txt": str(f.get("cantidad_txt", "") or "").strip(),
            "destino": str(f.get("destino", "") or "").strip(),
        })
    return limpias


# ------------------------------------------------ respaldo: lectura por columnas

# Nombres de pestaña que son rubros, no lugares: no generan un presupuesto aparte.
RUBROS = {"VIVERES", "VIVERE", "ALMACEN", "LIMPIEZA", "CARNE", "CARNICERIA",
          "VERDULERIA", "FRUTAS", "VERDURAS", "BEBIDAS", "PANADERIA", "GENERAL",
          "FIAMBRERIA", "LACTEOS", "CONGELADOS", "PEDIDO", "LISTA", "HOJA1",
          "PRODUCTOS", "TOTALES", "RESUMEN", "PRECIOS"}


def _es_rubro(nombre):
    """
    Una pestaña es un rubro solo si TODAS sus palabras son de rubro.

    "VIVERES" es un rubro; "VIVERES GUARDIAS" no, porque lo que la identifica
    es GUARDIAS, que es un destino. Con la regla laxa (cualquier palabra) se
    perdian pestañas como "Pedido Sojares" o "Viveres Guardias", que son
    entregas distintas y tienen que salir en presupuestos separados.
    """
    palabras = [p for p in norm(nombre).split() if len(p) > 2 and p != "DE"]
    if not palabras:
        return True
    return all(any(r == p or (len(p) > 4 and r in p) for r in RUBROS)
               for p in palabras)


def _leer_tabla(df):
    """
    Respaldo sin API: identifica que columna tiene los productos y cual las
    cantidades, y lee fila por fila. Sirve para Excel y CSV prolijos.
    """
    def _txt(c):
        t = "" if c is None else str(c).strip()
        return "" if t.lower() in ("nan", "none", "nat") else t

    df = df.map(_txt)
    if df.empty:
        return []

    col_texto, mejor = None, -1.0
    candidatas = []
    for c in df.columns:
        vals = [v for v in df[c].tolist() if v]
        if len(vals) < 2:
            continue
        numericos = sum(1 for v in vals if re.match(r"^[\d\-]", v)) / len(vals)
        if numericos >= 0.5:
            nums = []
            for v in vals:
                m = re.match(r"^([\d]+(?:[.,]\d+)?)", v.replace(".", "").replace(",", "."))
                if m:
                    try:
                        nums.append(float(m.group(1)))
                    except ValueError:
                        pass
            if nums and len(set(nums)) >= 3 and \
               max(nums.count(x) for x in set(nums)) / len(nums) <= 0.8:
                candidatas.append((sorted(nums)[len(nums) // 2], -numericos, c))
            continue
        variedad = len(set(norm(v) for v in vals)) / len(vals)
        largo = sum(len(v) for v in vals) / len(vals)
        puntaje = variedad * min(largo / 10, 1.5)
        if puntaje > mejor:
            col_texto, mejor = c, puntaje

    if col_texto is None:
        return []
    col_cant = min(candidatas)[2] if candidatas else None

    col_uni = None
    if col_cant is not None:
        for c in df.columns:
            if c in (col_cant, col_texto):
                continue
            vals = [v for v in df[c].tolist() if v]
            if vals and sum(1 for v in vals
                            if norm(v).rstrip(".") in UNIDADES
                            or norm(v).rstrip(".") in BULTOS) / len(vals) > 0.5:
                col_uni = c
                break

    filas = []
    for _, fila in df.iterrows():
        texto = fila[col_texto]
        if not texto or len(texto) < 3:
            continue
        cant = fila[col_cant] if col_cant is not None else ""
        if col_uni is not None and fila[col_uni]:
            cant = f"{cant} {fila[col_uni]}".strip()
        filas.append({"texto": texto, "cantidad_txt": cant, "destino": ""})
    return _descartar_encabezados(filas)


def _descartar_encabezados(filas):
    basura = {"PRODUCTO", "ITEM", "DESCRIPCION", "DESCRIPCION DEL PROD.", "CANTIDAD",
              "UNIDAD", "CATEGORIA", "TOTAL", "SUBTOTAL", "PRECIO", "PRECIO UNIT",
              "OBSERVACION", "TOTALES", "PRESUPUESTO TOTAL"}
    return [f for f in filas if norm(f["texto"]) not in basura]


def _leer_pdf(ruta):
    import pdfplumber
    filas = []
    with pdfplumber.open(ruta) as pdf:
        for pagina in pdf.pages:
            for tabla in (pagina.extract_tables() or []):
                filas += _leer_tabla(pd.DataFrame(tabla))
            if not filas:
                for linea in (pagina.extract_text() or "").split("\n"):
                    m = re.match(r"(.+?)\s+([\d.,]+\s*\w*)$", linea.strip())
                    if m and len(m.group(1)) > 3:
                        filas.append({"texto": m.group(1),
                                      "cantidad_txt": m.group(2), "destino": ""})
    return _descartar_encabezados(filas)


def _leer_por_columnas(ruta):
    """Sin API: sirve para Excel y CSV prolijos, con una columna de cada cosa."""
    ext = os.path.splitext(ruta)[1].lower()
    if ext in (".xlsx", ".xls", ".xlsm"):
        return _leer_tabla(pd.read_excel(ruta, header=None, dtype=str))
    if ext in (".csv", ".tsv"):
        sep = "\t" if ext == ".tsv" else None
        return _leer_tabla(pd.read_csv(ruta, header=None, dtype=str, sep=sep, engine="python"))
    if ext == ".pdf":
        return _leer_pdf(ruta)
    raise ValueError(f"No se pudo leer el archivo {ext} sin ANTHROPIC_API_KEY")


# ---------------------------------------------------------------- diccionario

_IDF_CACHE = {}


def _tokens(texto):
    """Palabras utiles del pedido, con las medidas normalizadas ('500 grs' -> '500GR')."""
    t = norm(texto).replace("GRS", "GR").replace("GRAMOS", "GR").replace("LTS", "LT")
    t = re.sub(r"(\d+)\s*(GR|KG|ML|LT|L|CC|U)\b", r"\1\2", t)
    return [p for p in t.split() if len(p) > 2 and p not in PALABRAS_VACIAS]


def _cargar_idf():
    """
    Peso de cada palabra segun lo rara que sea en el catalogo.

    "JABON" aparece en cientos de productos y no distingue nada; "HELLMANNS"
    aparece en tres y alcanza sola para identificar el producto. Sin esto,
    buscar "Jabon Dove tocador" devolvia el jabon de tocador mas vendido en vez
    del Dove.
    """
    if _IDF_CACHE:
        return _IDF_CACHE
    import math
    from collections import Counter
    filas = consultar("SELECT descripcion FROM v_catalogo_presupuestos")
    total = max(len(filas), 1)
    frecuencia = Counter()
    for f in filas:
        frecuencia.update(set(norm(f["descripcion"]).split()))
    for palabra, veces in frecuencia.items():
        _IDF_CACHE[palabra] = math.log(total / (1 + veces))
    return _IDF_CACHE


# Puntaje minimo para aceptar un producto encontrado por nombre. Calibrado
# contra los pedidos reales: mas bajo trae ruido, mas alto pierde marcas.
UMBRAL_CATALOGO = 4.0


def buscar_en_catalogo(texto, limite=4):
    """
    Ultimo recurso cuando el termino no esta en el diccionario.

    Los clientes escriben la marca primero ("Kechut Hellmanns de 500grs",
    "Jabon Omo 800 gr") y el diccionario esta armado por termino generico, asi
    que esos casos caian en SIN MATCH aunque el producto existiera.
    """
    palabras = [p for p in _tokens(texto) if not re.fullmatch(r"[\d.,]+", p)]
    if not palabras:
        return []

    idf = _cargar_idf()
    condiciones = " OR ".join(["descripcion LIKE %s"] * len(palabras))
    filas = consultar(
        f"SELECT idproducto AS codprod, descripcion AS producto "
        f"FROM v_catalogo_presupuestos "
        f"WHERE ({condiciones}) AND precio_venta > 0 LIMIT 400",
        tuple(f"%{p}%" for p in palabras),
    )
    if not filas:
        return []

    puntuados = []
    for f in filas:
        nombre = set(norm(f["producto"]).split())
        puntaje = sum(idf.get(p, 6.0) for p in palabras if p in nombre)
        if puntaje >= UMBRAL_CATALOGO:
            puntuados.append((puntaje, f))
    puntuados.sort(key=lambda x: -x[0])
    return [{"codprod": f["codprod"], "producto": f["producto"]}
            for _, f in puntuados[:limite]]


def cargar_diccionario():
    dicc = consultar(
        "SELECT termino, codprod, producto, prioridad "
        "FROM presupuesto_diccionario ORDER BY termino, prioridad")
    sinon = {norm(r["escribe"]): norm(r["buscar"])
             for r in consultar("SELECT escribe, buscar FROM presupuesto_sinonimos")}
    empaq = consultar("SELECT rubro, unidad_cliente, equivale_a FROM presupuesto_empaques")
    negro = {r["codprod"] for r in consultar("SELECT codprod FROM presupuesto_no_cotizar")}

    porTermino = {}
    for r in dicc:
        if r["codprod"] in negro:
            continue
        porTermino.setdefault(norm(r["termino"]), []).append(r)
    return porTermino, sinon, empaq, negro


def historial_cliente(ruc, meses=MESES_HISTORIAL):
    """{codprod: facturas} de lo que ese RUC compro."""
    if not ruc:
        return {}
    filas = consultar(
        "SELECT Codprod AS codprod, COUNT(DISTINCT Factura) AS facturas "
        "FROM ventas WHERE ruc = %s "
        "  AND Fecha_Hora >= DATE_SUB(CURDATE(), INTERVAL %s MONTH) "
        "GROUP BY Codprod", (ruc, meses))
    return {r["codprod"]: r["facturas"] for r in filas}


def historial_referencia(nombre):
    """Historial de otro cliente, buscado por razon social."""
    if not nombre:
        return {}
    filas = consultar(
        "SELECT Codprod AS codprod, COUNT(DISTINCT Factura) AS facturas "
        "FROM ventas WHERE razon_social LIKE %s "
        "  AND Fecha_Hora >= DATE_SUB(CURDATE(), INTERVAL %s MONTH) "
        "GROUP BY Codprod", (f"%{nombre}%", MESES_HISTORIAL))
    return {r["codprod"]: r["facturas"] for r in filas}


def precios(codprods):
    if not codprods:
        return {}
    marcas = ",".join(["%s"] * len(codprods))
    filas = consultar(
        f"SELECT idproducto AS codprod, precio_venta, ultimo_costo_compra, stock_teorico "
        f"FROM v_catalogo_presupuestos WHERE idproducto IN ({marcas})",
        tuple(codprods))
    return {r["codprod"]: r for r in filas}


# ---------------------------------------------------------------- resolucion

def buscar_termino(texto, porTermino, sinon):
    """
    El termino literal manda sobre el sinonimo.

    Si existe VIRULANA en el diccionario y ademas el sinonimo
    VIRULANA -> LANA DE ACERO, aplicar el sinonimo primero desvia hacia un
    termino que no existe y el producto queda sin resolver. Un sinonimo tiene
    que ser un atajo cuando falta el termino, nunca un desvio del que si esta.
    """
    t = norm(texto)
    if t in porTermino:
        return t
    t = sinon.get(t, t)
    if t in porTermino:
        return t
    # coincidencia parcial: el termino mas largo que aparezca en el texto
    candidatos = [k for k in porTermino if k in t or t in k]
    return max(candidatos, key=len) if candidatos else None


def equivalencia_empaque(termino, unidad, empaq):
    """
    Busca cuantas unidades trae 1 caja/bolsa/fardo del rubro pedido.

    El match del rubro tiene que ser del TERMINO COMPLETO, no de la primera
    palabra: 'JABON EN PAN' y 'JABON DE TOCADOR' arrancan igual pero traen
    20 y 30 unidades por caja. Comparar solo la primera palabra hacia que el
    jabon en pan se cotizara x30 en vez de x20.
    """
    u = norm(unidad).rstrip("S")
    t = norm(termino)
    candidatos = []
    for e in empaq:
        if norm(e["unidad_cliente"]).rstrip("S") != u:
            continue
        r = norm(e["rubro"])
        if r == t:
            return float(e["equivale_a"])          # coincidencia exacta, gana siempre
        if r in t or t in r:
            candidatos.append((len(r), float(e["equivale_a"])))
    # entre parciales, el rubro mas especifico (mas largo)
    return max(candidatos)[1] if candidatos else None


def resolver(linea, porTermino, sinon, empaq, hist, hist_ref):
    """Devuelve el dict de una linea del presupuesto."""
    texto = linea["texto"]
    qty, uni = parse_cantidad(linea.get("cantidad_txt", ""))

    base = dict(pide=texto, cantidad_pedida=linea.get("cantidad_txt", ""),
                codprod=None, producto=None, cantidad=None,
                confianza=None, origen=None, alternativas="", aviso="")

    if qty is None:
        base.update(confianza="OMITIR", origen="sin cantidad",
                    aviso="el cliente no puso cantidad")
        return base

    termino = buscar_termino(texto, porTermino, sinon)
    if termino:
        opciones = porTermino[termino]
    else:
        # no esta en el diccionario: se busca directo en el catalogo
        opciones = buscar_en_catalogo(texto)
        if opciones:
            base["aviso"] = "encontrado por el nombre, no por el diccionario - VERIFICAR"
        if not opciones:
            base.update(confianza="SIN MATCH", origen="no esta en el catalogo",
                        aviso="cargar a mano")
            return base
        termino = norm(texto)

    # --- ordenar candidatos segun la fuente disponible
    def ordenar(h):
        con = [(o, h.get(o["codprod"], 0)) for o in opciones]
        con = [c for c in con if c[1] > 0]
        con.sort(key=lambda x: -x[1])
        return con

    del_catalogo = termino not in porTermino
    propios = ordenar(hist)
    if propios:
        total = sum(f for _, f in propios)
        concentracion = propios[0][1] / total if total else 0
        elegidas = [o for o, _ in propios]
        if concentracion >= UMBRAL_ALTA:
            conf, orig = "ALTA", f"lo compra siempre ({round(concentracion*100)}%)"
        else:
            conf, orig = "MEDIA", f"lo compra alternando ({round(concentracion*100)}%)"
    else:
        refs = ordenar(hist_ref)
        if refs:
            elegidas = [o for o, _ in refs]
            conf, orig = "BAJA", "referencia de otro cliente"
        else:
            elegidas = opciones
            conf, orig = "BAJA", ("encontrado en el catalogo por el nombre"
                                  if del_catalogo else "ranking general")

    # --- si el cliente nombro marca o variedad, respetarla
    # "Yerba kurupi" no es lo mismo que "Yerba Colon", y "fideo tallarin" no es
    # "fideo spaghetti". Se buscan las palabras del pedido que NO son el termino
    # generico y se prioriza el candidato que las tenga.
    palabras = {p for p in norm(texto).split()
                if len(p) > 3 and p not in norm(termino).split()
                and p not in PALABRAS_VACIAS}
    if palabras:
        con_marca = [o for o in elegidas
                     if palabras & set(norm(o["producto"]).split())]
        if con_marca and con_marca != elegidas[:len(con_marca)]:
            resto = [o for o in elegidas if o not in con_marca]
            elegidas = con_marca + resto
            orig += " / respeta lo que pidio"

    # --- si el cliente escribio el tamano DENTRO del texto, ese manda
    # "SAL FINA DE 5KG" cotizaba el envase de 1KG (2.500 en vez de 13.000) y
    # "LEVADURA SECA DE 500GRS" el de 50GR. El tamano estaba escrito y no se
    # usaba: solo se miraba la columna de cantidad.
    tam_pedido = tamano_producto(texto)
    if tam_pedido[0]:
        coinciden = [o for o in elegidas
                     if _mismo_tamano(tamano_producto(o["producto"]), tam_pedido)]
        if not coinciden:
            # ninguna opcion del diccionario tiene ese tamano: se busca en el
            # catalogo entero antes de cotizar el envase equivocado
            for o in buscar_en_catalogo(texto, limite=8):
                if _mismo_tamano(tamano_producto(o["producto"]), tam_pedido):
                    coinciden.append(o)
        if coinciden:
            elegidas = coinciden + [o for o in elegidas if o not in coinciden]
            orig += f" / {tam_pedido[0]:g}{tam_pedido[1]} como pidio"
        else:
            base["aviso"] = (
                f"el cliente pidio {tam_pedido[0]:g}{tam_pedido[1]} y no hay "
                f"esa presentacion - VERIFICAR")

    # --- filtrar por unidad/tamano que pidio el cliente
    if uni == "U" and qty:
        # El cliente pidio unidades sueltas: hay que preferir el producto que
        # se vende por unidad. Sin esto, 'ajo 10 unidades' caia en AJO POR KG
        # y cotizaba 10 kilos de ajo en vez de 10 cabezas.
        sueltos = [o for o in elegidas
                   if norm(o["producto"]).endswith("POR UNIDAD")
                   or "X UNIDAD" in norm(o["producto"])
                   or "POR UNID" in norm(o["producto"])]
        granel = [o for o in elegidas if "POR KG" in norm(o["producto"])]
        if sueltos:
            o = sueltos[0]
            base.update(codprod=o["codprod"], producto=o["producto"], cantidad=qty,
                        origen=f"{orig} / por unidad")
        else:
            # sin version por unidad: se usa el primero, pero avisando, porque
            # si es un producto a granel la cantidad queda mal.
            o = elegidas[0]
            base.update(codprod=o["codprod"], producto=o["producto"], cantidad=qty,
                        origen=orig)
            if o in granel:
                base["aviso"] = "el cliente pidio unidades y esto se vende por kilo - VERIFICAR"
    elif uni in ("KG", "L") and qty:
        exactos = []
        for o in elegidas:
            v, u = tamano_producto(o["producto"])
            if v and u == uni and qty % v == 0:
                exactos.append((o, v))
        if exactos:
            o, v = max(exactos, key=lambda x: x[1])
            base.update(codprod=o["codprod"], producto=o["producto"],
                        cantidad=int(qty / v), origen=f"{orig} / {int(qty/v)} x {v:g}{u}")
        else:
            granel = [o for o in elegidas if "POR KG" in norm(o["producto"])]
            if granel:
                o = granel[0]
                base.update(codprod=o["codprod"], producto=o["producto"], cantidad=qty,
                            origen=f"{orig} / a granel")
            else:
                # Ninguna presentacion divide exacto. Se toma la mas grande que
                # entre y el resto se avisa: 8 kilos con bolsas de 5 son 1 bolsa
                # y sobran 3, no 8 bolsas de 5 (que serian 40 kilos).
                caben = []
                for o in elegidas:
                    v, u = tamano_producto(o["producto"])
                    if v and u == uni and v <= qty:
                        caben.append((o, v))
                if caben:
                    o, v = max(caben, key=lambda x: x[1])
                    entero = int(qty // v)
                    resto = qty - entero * v
                    base.update(codprod=o["codprod"], producto=o["producto"],
                                cantidad=entero,
                                origen=f"{orig} / {entero} x {v:g}{u}")
                    base["aviso"] = (
                        f"faltan {resto:g}{u}: completar con otra presentacion")
                else:
                    o = elegidas[0]
                    base.update(codprod=o["codprod"], producto=o["producto"],
                                cantidad=qty, origen=orig)
                    base["aviso"] = "no hay presentacion que de exacto - VERIFICAR"
    elif uni and norm(uni) in BULTOS:
        factor = equivalencia_empaque(termino, uni, empaq)
        o = elegidas[0]
        # Si el producto elegido YA SE VENDE POR BULTO, aplicar ademas la
        # equivalencia multiplica dos veces.
        #
        # No alcanza con comparar la misma palabra: el cliente pide "6 cajas de
        # pollo" y el producto es "POLLO ENTERO CONG PECHUGON POR BOLSA".
        # Palabras distintas, mismo bulto. Sin esto, 6 cajas se convertian en
        # 90 bolsas: 24 millones de mas en un solo renglon.
        #
        # Esta regla le gana a la equivalencia declarada: "PAPA bolsa = 18
        # kilos" vale cuando el producto se vende POR KG, no cuando ya es
        # "PAPA NEGRA POR BOLSA". Si el producto viene por bulto, 3 bolsas
        # son 3 bolsas y punto.
        u_simple = norm(uni).rstrip("S")
        nombre_prod = norm(o["producto"])
        if u_simple:
            palabras_bulto = ("BOLSA", "CAJA", "FARDO", "PLANCHA", "PAQUETE",
                              "BIDON", "TIRA", "PACK", "FUNDA")
            if any(f"POR {b}" in nombre_prod or f"X {b}" in nombre_prod
                   or nombre_prod.endswith(b) for b in palabras_bulto):
                base["origen"] = f"{orig} / el producto ya se vende por bulto"
                factor = None
        if factor:
            base.update(codprod=o["codprod"], producto=o["producto"],
                        cantidad=qty * factor,
                        origen=f"{orig} / {qty:g} {uni.lower()} x {factor:g}")
        elif norm(uni) in BULTOS_AMBIGUOS or base.get("origen"):
            base.update(codprod=o["codprod"], producto=o["producto"], cantidad=qty,
                        origen=base.get("origen") or f"{orig} / 1 {uni.lower()} = 1 unidad")
        else:
            base.update(codprod=o["codprod"], producto=o["producto"], cantidad=qty,
                        origen=orig, aviso=f"no se cuantas unidades trae 1 {uni.lower()}")
            conf = "MEDIA" if conf == "ALTA" else conf
    else:
        o = elegidas[0]
        base.update(codprod=o["codprod"], producto=o["producto"], cantidad=qty, origen=orig)

    base["confianza"] = conf
    base["alternativas"] = " | ".join(
        f"[{o['codprod']}] {o['producto']}" for o in elegidas[1:4])
    return base


# ---------------------------------------------------------------- salida

def guardar_entrada(nombre, contenido_bytes):
    """Guarda el archivo que subio el usuario en presupuestos/entrada y devuelve la ruta."""
    seguro = re.sub(r"[^A-Za-z0-9._-]", "_", os.path.basename(nombre))
    destino = os.path.join(DIR_ENTRADA, f"{datetime.now():%Y%m%d_%H%M%S}_{seguro}")
    with open(destino, "wb") as f:
        f.write(contenido_bytes)
    return destino


def generar(ruta_archivo, ruc=None, referencia=None, salida=None):
    """
    Procesa el pedido y devuelve la ruta de UN solo Excel.

    Si el pedido viene separado por destino (estancias, retiros, residencias),
    cada uno va en su propia pestaña del mismo archivo, mas una hoja RESUMEN
    con los totales. Antes salia un archivo por destino y comparar precios
    entre ellos obligaba a abrir seis ventanas.
    """
    print(f"Leyendo {os.path.basename(ruta_archivo)} ...")
    lineas = leer_archivo(ruta_archivo)
    print(f"  {len(lineas)} lineas detectadas")

    porTermino, sinon, empaq, negro = cargar_diccionario()
    hist = historial_cliente(ruc)
    hist_ref = historial_referencia(referencia)
    if ruc:
        print(f"  historial del RUC {ruc}: {len(hist)} productos")
    if referencia:
        print(f"  referencia {referencia}: {len(hist_ref)} productos")

    # Las lineas sin destino van a un presupuesto "General": si se las filtra
    # por no coincidir con ningun destino con nombre, desaparecen sin que nadie
    # se entere, que es la peor forma de fallar.
    con_nombre = sorted({(l.get("destino") or "").strip() for l in lineas} - {""})
    sueltas = [l for l in lineas if not (l.get("destino") or "").strip()]
    if con_nombre and sueltas:
        for l in sueltas:
            l["destino"] = "General"
        con_nombre.append("General")
    destinos = con_nombre or [""]
    if len(destinos) > 1:
        print(f"  el pedido viene separado en {len(destinos)} destinos: "
              f"{', '.join(destinos)}")

    ORDEN = ["pide", "cantidad_pedida", "confianza", "codprod", "producto", "cantidad",
             "precio", "subtotal", "costo", "stock", "origen", "aviso", "alternativas"]

    hojas = []          # [(nombre_hoja, destino, df)]
    for destino in destinos:
        del_destino = [l for l in lineas
                       if (l.get("destino") or "").strip() == destino] if destino else lineas
        if not del_destino:
            continue
        filas = [resolver(l, porTermino, sinon, empaq, hist, hist_ref) for l in del_destino]
        _agregar_precios(filas)
        df = pd.DataFrame(filas).reindex(columns=ORDEN)
        hojas.append((_nombre_hoja(destino, [h[0] for h in hojas]), destino, df))

    if not salida:
        etiqueta = re.sub(r"[^A-Za-z0-9_-]", "_",
                          os.path.splitext(os.path.basename(ruta_archivo))[0])[:40]
        salida = os.path.join(DIR_SALIDA,
                              f"presupuesto_{etiqueta}_{datetime.now():%Y%m%d_%H%M}.xlsx")

    _escribir_libro(hojas, salida, ruc=ruc, origen=os.path.basename(ruta_archivo))

    for nombre_hoja, destino, df in hojas:
        print(f"\n  [{destino or 'unico'}]")
        for nivel, ayuda in (("ALTA", "precargado, revisar por arriba"),
                             ("MEDIA", "elegir entre las alternativas"),
                             ("BAJA", "cliente nuevo o producto nuevo"),
                             ("SIN MATCH", "cargar a mano"),
                             ("OMITIR", "el cliente no puso cantidad")):
            print(f"  {nivel:<10}{(df.confianza == nivel).sum():3d}  {ayuda}")
        total = df.subtotal.fillna(0).sum() if "subtotal" in df else 0
        print(f"  Total: Gs. {total:,.0f}".replace(",", "."))
    print(f"\n  Archivo: {salida}")
    return salida


def _nombre_hoja(destino, usados):
    """
    Nombre de pestaña valido para Excel: hasta 31 caracteres, sin  []:*?/\
    y sin repetirse.
    """
    base = re.sub(r"[\[\]:*?/\\]", "-", (destino or "Presupuesto").strip())
    base = base.replace("·", "-").strip() or "Presupuesto"
    base = base[:31]
    nombre, n = base, 2
    while nombre in usados:
        corte = 31 - len(f" ({n})")
        nombre = f"{base[:corte]} ({n})"
        n += 1
    return nombre


def _agregar_precios(filas):
    """Completa precio, costo, stock y las alertas de margen."""
    pr = precios([f["codprod"] for f in filas if f["codprod"]])
    for f in filas:
        p = pr.get(f["codprod"])
        if not p:
            continue
        f["precio"] = p["precio_venta"]
        f["costo"] = p["ultimo_costo_compra"]
        f["stock"] = p["stock_teorico"]
        f["subtotal"] = (f["cantidad"] or 0) * (p["precio_venta"] or 0)
        pv, cc = p["precio_venta"] or 0, p["ultimo_costo_compra"] or 0
        if pv and cc and pv < cc:
            f["aviso"] = (f["aviso"] + " | " if f["aviso"] else "") + "PERDIDA: precio bajo el costo"
        elif pv and cc and pv < cc * 1.05:
            f["aviso"] = (f["aviso"] + " | " if f["aviso"] else "") + "margen menor al 5%"
        elif not pv:
            f["aviso"] = (f["aviso"] + " | " if f["aviso"] else "") + "sin precio cargado"


def _escribir_libro(hojas, salida, ruc=None, origen=""):
    """
    Un solo Excel con una pestaña por destino y una hoja RESUMEN adelante.

    hojas: [(nombre_hoja, destino, df)]
    """
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    COLOR = {"ALTA": "D9EAD3", "MEDIA": "FFF2CC", "BAJA": "FCE5CD",
             "SIN MATCH": "F4CCCC", "OMITIR": "EFEFEF"}
    MILES = "#,##0"
    ANCHO = {"pide": 26, "cantidad_pedida": 15, "confianza": 12, "codprod": 10,
             "producto": 42, "cantidad": 10, "precio": 13, "subtotal": 14,
             "costo": 12, "stock": 9, "origen": 34, "aviso": 36, "alternativas": 60}
    CABECERA = PatternFill("solid", fgColor="1F4E5F")

    with pd.ExcelWriter(salida, engine="openpyxl") as xw:
        if len(hojas) > 1:
            _hoja_resumen(xw, hojas, ruc, origen)

        for nombre_hoja, destino, df in hojas:
            df.to_excel(xw, index=False, sheet_name=nombre_hoja, startrow=2)
            ws = xw.sheets[nombre_hoja]

            titulo = f"PRESUPUESTO - {origen}"
            if destino:
                titulo += f"   |   DESTINO: {destino}"
            if ruc:
                titulo += f"   |   RUC {ruc}"
            ws.cell(1, 1, titulo).font = Font(bold=True, size=12)
            ws.cell(2, 1, f"Generado {datetime.now():%d/%m/%Y %H:%M} - "
                          f"REVISAR antes de emitir").font = Font(italic=True, size=9)

            cols = list(df.columns)
            for i, c in enumerate(cols, 1):
                cel = ws.cell(3, i)
                cel.font = Font(bold=True, color="FFFFFF")
                cel.fill = CABECERA
                cel.alignment = Alignment(horizontal="center", wrap_text=True)
                ws.column_dimensions[get_column_letter(i)].width = ANCHO.get(c, 16)

            idx_conf = cols.index("confianza") + 1 if "confianza" in cols else None
            numericas = [cols.index(c) + 1 for c in
                         ("precio", "subtotal", "costo", "cantidad", "stock")
                         if c in cols]

            for fila in range(4, len(df) + 4):
                conf = ws.cell(fila, idx_conf).value if idx_conf else None
                if conf in COLOR:
                    relleno = PatternFill("solid", fgColor=COLOR[conf])
                    for col in range(1, len(cols) + 1):
                        ws.cell(fila, col).fill = relleno
                for col in numericas:
                    ws.cell(fila, col).number_format = MILES

            if "subtotal" in cols:
                f = len(df) + 5
                c_sub = cols.index("subtotal") + 1
                ws.cell(f, c_sub - 1, "TOTAL").font = Font(bold=True)
                cel = ws.cell(f, c_sub, float(df["subtotal"].fillna(0).sum()))
                cel.font = Font(bold=True, size=12)
                cel.number_format = MILES
                ws.cell(f, c_sub + 1,
                        "estimado - faltan las lineas sin resolver").font = \
                    Font(italic=True, size=9)

            ws.freeze_panes = "A4"
            ws.auto_filter.ref = f"A3:{get_column_letter(len(cols))}{len(df) + 3}"


def _hoja_resumen(xw, hojas, ruc, origen):
    """Primera pestaña: totales por destino, para comparar de un vistazo."""
    from openpyxl.styles import Alignment, Font, PatternFill

    ws = xw.book.create_sheet("RESUMEN", 0)
    ws.cell(1, 1, f"RESUMEN - {origen}").font = Font(bold=True, size=13)
    if ruc:
        ws.cell(2, 1, f"RUC {ruc}").font = Font(italic=True, size=9)
    ws.cell(3, 1, f"Generado {datetime.now():%d/%m/%Y %H:%M}").font = \
        Font(italic=True, size=9)

    encabezados = ["DESTINO", "LINEAS", "ALTA", "MEDIA", "BAJA",
                   "SIN MATCH", "CON AVISO", "TOTAL Gs."]
    anchos = [34, 9, 8, 9, 8, 12, 12, 16]
    for i, (c, w) in enumerate(zip(encabezados, anchos), 1):
        cel = ws.cell(5, i, c)
        cel.font = Font(bold=True, color="FFFFFF")
        cel.fill = PatternFill("solid", fgColor="1F4E5F")
        cel.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[chr(64 + i)].width = w

    fila = 6
    for _, destino, df in hojas:
        avisos = int((df["aviso"].fillna("") != "").sum()) if "aviso" in df else 0
        valores = [destino or "Presupuesto", len(df),
                   int((df.confianza == "ALTA").sum()),
                   int((df.confianza == "MEDIA").sum()),
                   int((df.confianza == "BAJA").sum()),
                   int((df.confianza == "SIN MATCH").sum()),
                   avisos,
                   float(df["subtotal"].fillna(0).sum()) if "subtotal" in df else 0]
        for i, v in enumerate(valores, 1):
            cel = ws.cell(fila, i, v)
            if i == 8:
                cel.number_format = "#,##0"
        fila += 1

    ws.cell(fila + 1, 1, "TOTAL GENERAL").font = Font(bold=True, size=12)
    cel = ws.cell(fila + 1, 8,
                  float(sum(df["subtotal"].fillna(0).sum() for _, _, df in hojas)))
    cel.font = Font(bold=True, size=13)
    cel.number_format = "#,##0"
    ws.cell(fila + 3, 1,
            "Cada destino tiene su propia pestaña. Estimado: no incluye las "
            "líneas sin resolver.").font = Font(italic=True, size=9)


if __name__ == "__main__":
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    if not args:
        print(__doc__)
        sys.exit(0)

    def opt(nombre):
        if nombre in sys.argv:
            i = sys.argv.index(nombre)
            if i + 1 < len(sys.argv):
                return sys.argv[i + 1]
        return None

    generar(args[0], ruc=opt("--ruc"), referencia=opt("--referencia"),
            salida=opt("--salida"))
